# -*- coding: utf-8 -*-
"""
/***************************************************************************
 ODKGeo_QgisWktDialog
                                 A QGIS plugin
 This plugin converts ODK geo coordinate values to QGIS (flipped) WKT values.
 It processes points, traces (lines), and polygons.
 ***************************************************************************/
"""

import os
import gc
from openpyxl import load_workbook
from qgis.PyQt import uic
from qgis.PyQt import QtWidgets
from qgis.PyQt.QtWidgets import QMessageBox
from shapely.geometry import Point, LineString, Polygon

# Load UI file
FORM_CLASS, _ = uic.loadUiType(os.path.join(
    os.path.dirname(__file__), 'odk_geo_qgis_wkt_dialog_base.ui'))

class ODKGeo_QgisWktDialog(QtWidgets.QDialog, FORM_CLASS):
    def __init__(self, parent=None):
        """Constructor - Initializes the UI and connects signals to slots."""
        super(ODKGeo_QgisWktDialog, self).__init__(parent)
        self.setupUi(self)

        # Ensure convertButton exists in the UI
        if hasattr(self, "convertButton"):
            self.convertButton.clicked.connect(self.convert_coordinates)
        else:
            raise AttributeError("convertButton is missing from the UI file. Check your .ui design.")

        # Connect file and dropdown widgets to corresponding functions
        self.xlsFileWidget.fileChanged.connect(self.load_sheets)
        self.sheetDropdown.currentIndexChanged.connect(self.load_columns)

        # Add autoSelectCheckbox functionality if it exists in the UI (no additional features added for backward compatibility)
        if hasattr(self, 'autoSelectCheckbox'):
            self.autoSelectCheckbox.stateChanged.connect(self.load_columns)

    def load_sheets(self):
        """Load sheets from the selected .xlsx file into the sheet dropdown and auto-load first sheet's columns."""
        file_path = self.xlsFileWidget.filePath()
        if not file_path.endswith(".xlsx"):
            QMessageBox.warning(self, "Invalid File", "Please select a valid .xlsx file.")
            return

        try:
            workbook = load_workbook(file_path, read_only=False)  # Open in normal mode for writing
            self.sheetDropdown.clear()
            self.sheetDropdown.addItems(workbook.sheetnames)
            self.workbook = workbook

            # Auto-select the first sheet and load columns immediately
            if workbook.sheetnames:
                self.sheetDropdown.setCurrentIndex(0)  # Select first sheet
                self.load_columns()  # Populate columns automatically

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to read .xlsx file: {e}")

    def load_columns(self):
        """Loads column headers from the selected sheet into dropdowns and auto-selects specific columns if enabled."""
        selected_sheet = self.sheetDropdown.currentText()
        if not selected_sheet or not hasattr(self, 'workbook'):
            return

        try:
            # Get the headers from the first row
            sheet = self.workbook[selected_sheet]
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)) if cell.value]

            # Populate dropdowns with column names
            self.pointColumnDropdown.clear()
            self.traceColumnDropdown.clear()
            self.polygonColumnDropdown.clear()

            self.pointColumnDropdown.addItems(headers)
            self.traceColumnDropdown.addItems(headers)
            self.polygonColumnDropdown.addItems(headers)

            # If auto-select is enabled, check for missing columns
            if hasattr(self, 'autoSelectCheckbox') and self.autoSelectCheckbox.isChecked():
                odk_geopoint_col = "Single point site coordinates (latitude and longitude)."
                odk_geotrace_col = "Record surface feature as line."
                odk_geoshape_col = "Record surface feature as polygon."
                
                missing_columns = []
                if odk_geopoint_col in headers:
                    self.pointColumnDropdown.setCurrentText(odk_geopoint_col)
                else:
                    missing_columns.append(odk_geopoint_col)

                if odk_geotrace_col in headers:
                    self.traceColumnDropdown.setCurrentText(odk_geotrace_col)
                else:
                    missing_columns.append(odk_geotrace_col)

                if odk_geoshape_col in headers:
                    self.polygonColumnDropdown.setCurrentText(odk_geoshape_col)
                else:
                    missing_columns.append(odk_geoshape_col)

                # If any required column is missing, show an error
                if missing_columns:
                    QMessageBox.critical(
                        self, 
                        "Missing Required Columns", 
                        f"The following required columns are missing:\n\n- " + "\n- ".join(missing_columns) + 
                        "\n\nPlease select a sheet that contains these columns or disable auto-select."
                    )
                    return  # Stop execution

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load columns: {e}")

    def convert_coordinates(self):
        """Converts selected coordinate columns into WKT format (Point, LineString, Polygon)."""
        selected_sheet = self.sheetDropdown.currentText()
        point_column = self.pointColumnDropdown.currentText()
        trace_column = self.traceColumnDropdown.currentText()
        polygon_column = self.polygonColumnDropdown.currentText()

        # Get user-defined trace result column name
        user_trace_column_name = self.traceResultColumnName.text().strip()
        if not user_trace_column_name:  # If empty, use default
            user_trace_column_name = "QGIS Trace WKT"

        # Get user-defined trace result column name
        user_poly_column_name = self.polyResultColumnName.text().strip()
        if not user_poly_column_name:  # If empty, use default
            user_poly_column_name = "QGIS Poly WKT"

        if not selected_sheet or not hasattr(self, 'workbook'):
            QMessageBox.warning(self, "Error", "No sheet selected or workbook not loaded.")
            return

        file_path = self.xlsFileWidget.filePath()
        try:
            # Open workbook for editing
            workbook = load_workbook(file_path)
            sheet = workbook[selected_sheet]

            # Get existing headers and determine column positions
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            existing_columns = {header: idx + 1 for idx, header in enumerate(headers) if header}

            # Determine column index for user-defined trace column
            trace_col_idx = existing_columns.get(user_trace_column_name, len(headers) + 1)

            
            # Determine column index for user-defined trace column
            poly_col_idx = existing_columns.get(user_poly_column_name, len(headers) + 2)

            # Add WKT column headers if it doesn’t exist
            if user_trace_column_name not in existing_columns:
                sheet.cell(row=1, column=trace_col_idx, value=user_trace_column_name)

            # Add WKT column headers if it doesn’t exist
            if user_poly_column_name not in existing_columns:
                sheet.cell(row=1, column=poly_col_idx, value=user_poly_column_name)

            # Iterate over rows and convert trace coordinates
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
                if trace_column and row[headers.index(trace_column)].value:
                    trace = [self.flip_coordinates(coord) for coord in row[headers.index(trace_column)].value.split(';')]
                    sheet.cell(row=row_index, column=trace_col_idx, value=LineString(trace).wkt)

            # Iterate over rows and convert trace coordinates
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
                if polygon_column and row[headers.index(polygon_column)].value:
                    poly = [self.flip_coordinates(coord) for coord in row[headers.index(polygon_column)].value.split(';')]
                    sheet.cell(row=row_index, column=poly_col_idx, value=Polygon(poly).wkt)

            # Save changes and clean up
            workbook.save(file_path)
            workbook.close()
            del workbook
            gc.collect()  # Free memory

            QMessageBox.information(self, "Success", f"Coordinates converted and saved to column: {user_trace_column_name}!")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to convert coordinates: {e}")


    def flip_coordinates(self, coordinate):
        """
        Flips ODK coordinates from (latitude, longitude) to (longitude, latitude).
        Expects coordinates in "lat lon" format.
        """
        coords = coordinate.split()
        if len(coords) >= 2:
            return float(coords[1]), float(coords[0])  # Swap lat and lon
        raise ValueError(f"Invalid coordinate format: {coordinate}")